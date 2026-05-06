"""Shared test fixtures: synthetic Seur invoice/workbook builders and
pointers to real-data fixtures."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import pytest
from openpyxl import Workbook

from courier_automation.parsers.correos import CORREOS_COLUMNS, CORREOS_RAW_COLUMNS
from courier_automation.parsers.seur import SEUR_COLUMNS
from courier_automation.parsers.ups import UPS_COLUMNS
from courier_automation.parsers.seitrans import SEITRANS_COLUMNS, SEITRANS_RAW_COLUMNS

FIXTURES_DIR = Path(__file__).parent / "fixtures"
SEUR_RAW_DIR = FIXTURES_DIR / "seur" / "raw"
SEUR_GOLDEN_DIR = FIXTURES_DIR / "seur" / "golden"
SEUR_SYNTHETIC_DIR = FIXTURES_DIR / "seur" / "synthetic"
SEITRANS_RAW_DIR = FIXTURES_DIR / "seitrans" / "raw"
SEITRANS_GOLDEN_DIR = FIXTURES_DIR / "seitrans" / "golden"
CORREOS_RAW_DIR = FIXTURES_DIR / "correos" / "raw"
CORREOS_GOLDEN_DIR = FIXTURES_DIR / "correos" / "golden"
UPS_RAW_DIR = FIXTURES_DIR / "ups" / "raw"
UPS_GOLDEN_DIR = FIXTURES_DIR / "ups" / "golden"


def _make_default_row(line_number: int = 1) -> dict[str, Any]:
    """Build a single Seur row with sensible defaults for every column.

    Tests can override individual fields by passing a dict to `make_seur_invoice`.
    Numeric columns get 0/0.0; date columns get a fixed date; text columns get
    a placeholder string.
    """
    row: dict[str, Any] = {
        "Codigo Cliente": "012345",
        "Serie Factura": "A",
        "Numero Factura": 289264,
        "Fecha Factura": datetime(2025, 4, 30),
        "Numero Linea": line_number,
        "Fecha Servicio": datetime(2025, 4, 15),
        "Salida / Entrada": "S",
        "Origen": "08",
        "Nombre Completo Origen": "Barcelona",
        "Destino": "28",
        "Nombre Completo Destino": "Madrid",
        "Servicio": "13",
        "Nombre Completo Servicio": "Servicio Estándar",
        "Producto": "1",
        "Nombre Completo Producto": "Producto Estándar",
        "U.A. Exp.": "U001",
        "Centro": "0289",
        "Numero Expedicion": f"E{line_number:08d}",
        "Fecha Exp.": datetime(2025, 4, 15),
        "Informacion Adicional": "",
        "Remitente": "Artero",
        "Direccion Remitente": "Calle Falsa 123",
        "Poblacion Remitente": "Barcelona",
        "C. Postal Remitente": "08001",
        "Destinatario": "Cliente",
        "Direccion Destinatario": "Avenida Real 456",
        "Poblacion Destinatario": "Madrid",
        "C. Postal Destinatario": "28001",
        "Referencia": f"REF{line_number:05d}",
        "Tipo Línea": "P",
        "Claves Expedicion": "",
        "Bultos": 1,
        "Peso": 1.5,
        "Peso Volumetrico": 1.0,
        "Ancho": 30.0,
        "Alto": 20.0,
        "Largo": 40.0,
        "Volumen": 0.024,
        "Clave Impuesto": "21",
        "Importe facturado (sin impuestos)": 5.50,
        "Valor Reembolso": 0.0,
        "Valor Asegurado": 0.0,
        "U.A. Consol.": "",
        "Codigo Cliente Consolidado": "",
        "Alias Razon Social CCC Consolidado": "",
        "Poliza flotante porte": 0.0,
        "Poliza flotante valor declarado": 0.0,
        "Portes": 4.50,
        "Reexpedicion Especial": 0.0,
        "Gestion Reembolso": 0.0,
        "Seguro": 0.0,
        "Cargo Combustible": 0.45,
        "Comprobante de entrega": 0.0,
        "Servicios Sabados": 0.0,
        "Sobrecargos No Encintable": 0.0,
        "Tasa Seguridad Int": 0.0,
        "Tasa Calidad del Dato": 0.0,
        "Tasa Islas": 0.0,
        "Tasa B2C": 0.55,
        "Tasa Cliente No Integrado": 0.0,
        "Suplemento Andorra": 0.0,
        "Zonas Remotas": 0.0,
        "Gestion Aduanas Salidas": 0.0,
        "Gestion Aduanas Llegadas": 0.0,
        "Suplidos": 0.0,
        "Aforos": 0.0,
        "Descuentos": 0.0,
        "Otros": 0.0,
    }
    assert set(row) == set(SEUR_COLUMNS), (
        f"default row schema drift: missing={set(SEUR_COLUMNS) - set(row)}, "
        f"extra={set(row) - set(SEUR_COLUMNS)}"
    )
    return row


def make_seur_invoice(
    path: Path,
    rows: Iterable[dict[str, Any]] | None = None,
    *,
    columns: tuple[str, ...] = SEUR_COLUMNS,
    sheet_name: str = "Sheet1",
) -> Path:
    """Write a Seur-shaped xlsx at `path` using openpyxl directly (so text cells
    stay text). If `rows` is None, writes one default row.
    """
    if rows is None:
        rows = [_make_default_row()]
    rows = list(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(columns))
    for row in rows:
        ws.append([row.get(col) for col in columns])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def make_empty_seur_workbook(path: Path) -> Path:
    """Build a workbook with a `Datos` sheet that has the 68 Seur headers and no
    data rows. Used as a writer-test target."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(list(SEUR_COLUMNS))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _make_default_seitrans_row(line_number: int = 1) -> dict[str, Any]:
    row: dict[str, Any] = {
        "CLIENTE_RAGIONE_SOCIALE": "Cliente Srl",
        # Real Seitrans: one DOCUMENTO_NUMERO per file, multiple SPEDIZIONE_NUMERO
        # rows underneath. Constant default mirrors that.
        "DOCUMENTO_NUMERO": 289264,
        "SPEDIZIONE_NUMERO": f"S{line_number:08d}",
        "MITTENTE_RAGIONE_SOCIALE": "Artero",
        "MITTENTE_NAZIONE_DESCRIZIONE": "Italia",
        "MITTENTE_CAP": "00100",
        "DESTINATARIO_RAGIONE_SOCIALE": "Cliente SL",
        "DESTINATARIO_LOCALITA": "Madrid",
        "DESTINATARIO_CAP": "28001",
        "DESTINATARIO_NAZIONE_DESCRIZIONE": "España",
        "IMBALLI": 1,
        "PESO_LORDO": 12.5,
        "VOLUME": 0.85,
        "PESO_TASSABILE": 12.5,
        "METRI_LINEARI": 0.0,
        "VOCE_DESCRIZIONE": "Documento",
        "IMPORTO_TOTALE_VALUTA": 120.0,
        "RIFERIMENTO_COMMITTENTE": f"REF{line_number:04d}",
        "RESA_DESCRIZIONE": "DAP",
        "SETTORE_DESCRIZIONE": "Logistica",
        "DOCUMENTO_DATA": datetime(2025, 4, 30),
    }
    assert set(row) == set(SEITRANS_RAW_COLUMNS), (
        f"default row schema drift: missing={set(SEITRANS_RAW_COLUMNS) - set(row)}, "
        f"extra={set(row) - set(SEITRANS_RAW_COLUMNS)}"
    )
    return row


def make_seitrans_invoice(
    path: Path,
    rows: Iterable[dict[str, Any]] | None = None,
    *,
    columns: tuple[str, ...] = SEITRANS_RAW_COLUMNS,
    sheet_name: str = "Risultato",
) -> Path:
    """Write a Seitrans-shaped xlsx at `path` using openpyxl directly."""
    if rows is None:
        rows = [_make_default_seitrans_row()]
    rows = list(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(columns))
    for row in rows:
        ws.append([row.get(col) for col in columns])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def make_empty_seitrans_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(list(SEITRANS_COLUMNS))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


@pytest.fixture
def seur_invoice_factory(tmp_path: Path):
    """Returns a callable that writes a synthetic Seur invoice xlsx into tmp_path."""

    def _factory(
        rows: Iterable[dict[str, Any]] | None = None,
        *,
        filename: str = "0289992025D0289264.xlsx",
        columns: tuple[str, ...] = SEUR_COLUMNS,
    ) -> Path:
        return make_seur_invoice(tmp_path / filename, rows, columns=columns)

    return _factory


@pytest.fixture
def empty_seur_workbook(tmp_path: Path) -> Path:
    return make_empty_seur_workbook(tmp_path / "seur_workbook.xlsx")


@pytest.fixture
def default_seur_row():
    return _make_default_row


@pytest.fixture
def seitrans_invoice_factory(tmp_path: Path):
    """Returns a callable that writes a synthetic Seitrans invoice xlsx into tmp_path."""

    def _factory(
        rows: Iterable[dict[str, Any]] | None = None,
        *,
        filename: str = "2025_04_30_INV0289264.xlsx",
        columns: tuple[str, ...] = SEITRANS_RAW_COLUMNS,
    ) -> Path:
        return make_seitrans_invoice(tmp_path / filename, rows, columns=columns)

    return _factory


@pytest.fixture
def empty_seitrans_workbook(tmp_path: Path) -> Path:
    return make_empty_seitrans_workbook(tmp_path / "seitrans_workbook.xlsx")


@pytest.fixture
def default_seitrans_row():
    return _make_default_seitrans_row


@pytest.fixture
def correos_invoice_factory(tmp_path: Path):
    """Returns a callable that writes a synthetic Correos invoice xlsx."""

    def _factory(
        rows: Iterable[dict[str, Any]] | None = None,
        *,
        filename: str = "2025_01_31 FAC_UNICO_F2501_14307.xlsx",
        columns: tuple[str, ...] = CORREOS_RAW_COLUMNS,
        invoice_number: str = "F250114307",
        invoice_date: datetime = datetime(2025, 1, 31),
    ) -> Path:
        return make_correos_invoice(
            tmp_path / filename,
            rows,
            columns=columns,
            invoice_number=invoice_number,
            invoice_date=invoice_date,
        )

    return _factory


@pytest.fixture
def empty_correos_workbook(tmp_path: Path) -> Path:
    return make_empty_correos_workbook(tmp_path / "correos_workbook.xlsx")


@pytest.fixture
def default_correos_row():
    return _make_default_correos_row


@pytest.fixture
def ups_invoice_factory(tmp_path: Path):
    """Returns a callable that writes a synthetic UPS billing CSV."""

    def _factory(
        rows: Iterable[dict[str, Any]] | None = None,
        *,
        filename: str = "Invoice_3961958_012225.csv",
        columns: tuple[str, ...] = UPS_COLUMNS,
    ) -> Path:
        return make_ups_invoice(tmp_path / filename, rows, columns=columns)

    return _factory


@pytest.fixture
def empty_ups_workbook(tmp_path: Path) -> Path:
    return make_empty_ups_workbook(tmp_path / "ups_workbook.xlsx")


@pytest.fixture
def default_ups_row():
    return _make_default_ups_row


def _make_default_correos_row(line_number: int = 1) -> dict[str, Any]:
    """Defaults for one shipment line in a synthetic Correos invoice."""
    row: dict[str, Any] = {
        "Nº ENVIO": f"E{line_number:09d}",
        "F.ALBARAN": datetime(2025, 1, 15),
        "F.ADMISION": datetime(2025, 1, 15),
        "REFERENCIA": f"REF{line_number:05d}",
        "Nº ENVIO CLIENTE": f"C{line_number:08d}",
        "BULTOS": 1,
        "PESO KILOS": 2.5,
        "VOLUMEN": 0.01,
        "C. LLAMADA": "",
        "PORTE": 4.50,
        "G. REEMBOLSO": 0.0,
        "DESEMBOLSO": 0.0,
        "REEXPEDICION": 0.0,
        "SEGUROS": 0.0,
        "SEGURO ESPECIAL": 0.0,
        "IMP. EXCESO MEDIDAS": 0.0,
        "SUPLEMENTO RECOGIDA": 0.0,
        "ENTREGA SABADO": 0.0,
        "SUPLEMTO O/D PORTUGAL": 0.0,
        "SUPLEMENTO DESTINO INGLATERRA": 0.0,
        "SUPLEMENTO COMBUSTIBLE": 0.45,
        "IMP. TOTAL": 5.99,
        "TIPO IMPOSITIVO": 21.0,
        "T. PORTE": 4.95,
        "PRODUCTO": "PAQ",
        "C. REMITENTE": "001",
        "N. REMITENTE": "Artero",
        "DOM. REMITENTE": "Calle Falsa 123",
        "POB. REMITENTE": "Barcelona",
        "C. P. REM.": "08001",
        "TEL. REMITENTE": "934567890",
        "C. C. REMITENTE": "001",
        "C. DESTINATARIO": "002",
        "N. DESTINATARIO": "Cliente",
        "DOM. DESTINATARIO": "Avenida Real 456",
        "POB. DESTINATARIO": "Madrid",
        "C. P. DESTINATARIO": "28001",
        "TEL. DESTINATARIO": "917654321",
        "C. C. DESTINATARIO": "001",
        "PLAZA ORIGEN": "08",
        "PLAZA DESTINO": "28",
        "PLAZA FACTURACION": "08",
        "V. ASEGURADO": 0.0,
        "IMP. REEMBOLSO": 0.0,
        "IMP. DESEMBOLSO": 0.0,
        "C. PAIS": "34",
        "OBSERVACIONES": "",
        "CLIENTE IMPUTACION": "1",
        "BAREMO": "STD",
        "F.ENTREGA": datetime(2025, 1, 16),
        "HORA ENTREGA": "10:30",
    }
    assert set(row) == set(CORREOS_RAW_COLUMNS), (
        f"default row schema drift: missing={set(CORREOS_RAW_COLUMNS) - set(row)}, "
        f"extra={set(row) - set(CORREOS_RAW_COLUMNS)}"
    )
    return row


def make_correos_invoice(
    path: Path,
    rows: Iterable[dict[str, Any]] | None = None,
    *,
    columns: tuple[str, ...] = CORREOS_RAW_COLUMNS,
    invoice_number: str = "F250114307",
    invoice_date: datetime = datetime(2025, 1, 31),
) -> Path:
    """Write a Correos-shaped xlsx with the inconvenient header band:
    row 0 = invoice-metadata labels, row 1 = invoice-metadata values,
    row 2 = shipment-line headers, rows 3+ = data."""
    if rows is None:
        rows = [_make_default_correos_row()]
    rows = list(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 0 — only the two labels the parser needs; pad to the line-table width.
    band_labels = ["Nº FACTURA", "F.FACTURA"] + [None] * (len(columns) - 2)
    ws.append(band_labels)
    band_values = [invoice_number, invoice_date] + [None] * (len(columns) - 2)
    ws.append(band_values)

    # Row 2 — actual shipment-line headers.
    ws.append(list(columns))
    for row in rows:
        ws.append([row.get(col) for col in columns])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def make_empty_correos_workbook(path: Path) -> Path:
    """Tiny workbook with a `Datos` sheet and the 58-col Correos header."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(list(CORREOS_COLUMNS))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _make_default_ups_row(line_number: int = 1) -> dict[str, Any]:
    """Defaults for one UPS billing-line row. UPS has 250 columns; we only
    set the handful the parser cares about (the no-null/plausibility set
    plus a couple of useful fields), the rest stay empty."""
    row: dict[str, Any] = dict.fromkeys(UPS_COLUMNS, "")
    row["Invoice Number"] = "000003961958"
    row["Invoice Date"] = "2025-01-22"
    row["Invoice Currency Code"] = "GBP"
    row["Invoice Amount"] = "-2.11"
    row["Tracking Number"] = f"1Z999AA1{line_number:010d}"
    row["Shipment Date"] = "2025-01-21"
    row["Charge Description"] = "Freight"
    row["Net Amount"] = "5.00"
    row["Entered Weight"] = "1.5"
    row["Billed Weight"] = "2.0"
    return row


def make_ups_invoice(
    path: Path,
    rows: Iterable[dict[str, Any]] | None = None,
    *,
    columns: tuple[str, ...] = UPS_COLUMNS,
) -> Path:
    """Write a headerless UPS-shaped CSV at `path`."""
    import csv

    if rows is None:
        rows = [_make_default_ups_row()]
    rows = list(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in rows:
            writer.writerow([row.get(col, "") for col in columns])
    return path


def make_empty_ups_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(list(UPS_COLUMNS))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _glob_xlsx(d: Path) -> list[Path]:
    return sorted(d.glob("*.xlsx")) if d.exists() else []


_REAL_SEUR_INVOICES = _glob_xlsx(SEUR_RAW_DIR)
_REAL_SEITRANS_INVOICES = _glob_xlsx(SEITRANS_RAW_DIR)
_REAL_CORREOS_INVOICES = _glob_xlsx(CORREOS_RAW_DIR)
_REAL_UPS_INVOICES = sorted(UPS_RAW_DIR.glob("*.csv")) if UPS_RAW_DIR.exists() else []


@pytest.fixture(
    params=_REAL_SEUR_INVOICES or [None],
    ids=lambda p: p.name if p else "no-fixtures",
)
def real_seur_invoice(request) -> Path:
    """Parametrized over every .xlsx in tests/fixtures/seur/raw/. When no
    fixtures are present, runs once and skips with a helpful message."""
    if request.param is None:
        pytest.skip(
            f"no real Seur fixtures at {SEUR_RAW_DIR} "
            "(copy a real invoice in to enable real-data tests)"
        )
    return request.param


@pytest.fixture(
    params=_REAL_SEITRANS_INVOICES or [None],
    ids=lambda p: p.name if p else "no-fixtures",
)
def real_seitrans_invoice(request) -> Path:
    """Parametrized over every .xlsx in tests/fixtures/seitrans/raw/."""
    if request.param is None:
        pytest.skip(
            f"no real Seitrans fixtures at {SEITRANS_RAW_DIR} "
            "(copy a real invoice in to enable real-data tests)"
        )
    return request.param


@pytest.fixture(
    params=_REAL_UPS_INVOICES or [None],
    ids=lambda p: p.name if p else "no-fixtures",
)
def real_ups_invoice(request) -> Path:
    """Parametrized over every .csv in tests/fixtures/ups/raw/."""
    if request.param is None:
        pytest.skip(
            f"no real UPS fixtures at {UPS_RAW_DIR} "
            "(copy a real invoice in to enable real-data tests)"
        )
    return request.param


@pytest.fixture(
    params=_REAL_CORREOS_INVOICES or [None],
    ids=lambda p: p.name if p else "no-fixtures",
)
def real_correos_invoice(request) -> Path:
    """Parametrized over every .xlsx in tests/fixtures/correos/raw/."""
    if request.param is None:
        pytest.skip(
            f"no real Correos fixtures at {CORREOS_RAW_DIR} "
            "(copy a real invoice in to enable real-data tests)"
        )
    return request.param
