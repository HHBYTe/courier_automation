"""End-to-end CLI tests using Typer's CliRunner."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from courier_automation.cli import (
    EXIT_LOCK,
    EXIT_MANIFEST_CONFLICT,
    EXIT_OK,
    EXIT_PLAUSIBILITY,
    EXIT_SCHEMA,
    EXIT_USAGE,
    app,
)
from courier_automation.parsers.seur import SEUR_COLUMNS
from courier_automation.store.workbook_appender import LOCK_SUFFIX
from tests.conftest import (
    make_empty_seur_workbook,
    make_empty_seitrans_workbook,
    make_seur_invoice,
    make_seitrans_invoice,
)

runner = CliRunner()


@pytest.fixture(autouse=True)
def isolated_manifest(tmp_path, monkeypatch):
    monkeypatch.setenv(
        "COURIER_AUTOMATION_MANIFEST", str(tmp_path / "manifest.sqlite")
    )


def _datos_row_count(path: Path) -> int:
    wb = load_workbook(path, read_only=True)
    try:
        return wb["Datos"].max_row - 1
    finally:
        wb.close()


def test_ingest_single_file_end_to_end(tmp_path, default_seur_row):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    invoice = make_seur_invoice(
        tmp_path / "0289992025D0289264.xlsx",
        rows=[default_seur_row(1), default_seur_row(2)],
    )

    result = runner.invoke(
        app,
        [
            "ingest",
            "seur",
            "--file",
            str(invoice),
            "--workbook",
            str(workbook),
        ],
    )
    assert result.exit_code == EXIT_OK, result.output
    assert "appended 2 rows" in result.output
    assert _datos_row_count(workbook) == 2


def test_ingest_is_idempotent(tmp_path, default_seur_row):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    invoice = make_seur_invoice(
        tmp_path / "0289992025D0289264.xlsx", rows=[default_seur_row(1)]
    )

    first = runner.invoke(
        app, ["ingest", "seur", "--file", str(invoice), "--workbook", str(workbook)]
    )
    assert first.exit_code == EXIT_OK
    assert _datos_row_count(workbook) == 1

    second = runner.invoke(
        app, ["ingest", "seur", "--file", str(invoice), "--workbook", str(workbook)]
    )
    assert second.exit_code == EXIT_OK
    assert "already ingested" in second.output
    assert _datos_row_count(workbook) == 1


def test_ingest_dry_run_writes_nothing(tmp_path, default_seur_row):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    invoice = make_seur_invoice(
        tmp_path / "0289992025D0289264.xlsx", rows=[default_seur_row(1)]
    )
    mtime_before = workbook.stat().st_mtime_ns

    result = runner.invoke(
        app,
        [
            "ingest",
            "seur",
            "--file",
            str(invoice),
            "--workbook",
            str(workbook),
            "--dry-run",
        ],
    )
    assert result.exit_code == EXIT_OK
    assert "would append 1 rows" in result.output
    assert "(dry-run)" in result.output
    assert workbook.stat().st_mtime_ns == mtime_before
    assert _datos_row_count(workbook) == 0


def test_ingest_month_processes_all_in_folder(tmp_path, default_seur_row):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    facturas = tmp_path / "Facturas" / "2025" / "04 - Abril"
    facturas.mkdir(parents=True)
    for n in (1, 2, 3):
        make_seur_invoice(
            facturas / f"0289992025D000000{n}.xlsx",
            rows=[default_seur_row(n)],
        )

    result = runner.invoke(
        app,
        [
            "ingest",
            "seur",
            "--month",
            "2025-04",
            "--folder",
            str(tmp_path / "Facturas"),
            "--workbook",
            str(workbook),
        ],
    )
    assert result.exit_code == EXIT_OK, result.output
    assert _datos_row_count(workbook) == 3
    assert "3 ingested" in result.output


def test_ingest_month_with_no_files_exits_usage(tmp_path):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    result = runner.invoke(
        app,
        [
            "ingest",
            "seur",
            "--month",
            "2025-04",
            "--folder",
            str(tmp_path / "Facturas"),
            "--workbook",
            str(workbook),
        ],
    )
    assert result.exit_code == EXIT_USAGE


def test_ingest_seitrans_single_file_end_to_end(tmp_path, default_seitrans_row):
    workbook = make_empty_seitrans_workbook(tmp_path / "wb.xlsx")
    invoice = make_seitrans_invoice(
        tmp_path / "2025_04_30_INV0289264.xlsx",
        rows=[default_seitrans_row(1)],
    )

    result = runner.invoke(
        app,
        [
            "ingest",
            "seitrans",
            "--file",
            str(invoice),
            "--workbook",
            str(workbook),
        ],
    )
    assert result.exit_code == EXIT_OK, result.output
    assert _datos_row_count(workbook) == 1


def test_ingest_seitrans_auto_discovers_latest_month(tmp_path, default_seitrans_row):
    """Latest month wins (April 2025), so only the two April files get
    ingested. Each file represents a distinct invoice, so we override
    DOCUMENTO_NUMERO per file — otherwise they'd all share the conftest
    default and trigger a manifest conflict."""
    workbook = make_empty_seitrans_workbook(tmp_path / "wb.xlsx")
    facturas = tmp_path / "Facturas" / "2025"
    facturas.mkdir(parents=True)

    def _row(line_number: int, doc_num: int):
        r = default_seitrans_row(line_number)
        r["DOCUMENTO_NUMERO"] = doc_num
        return r

    make_seitrans_invoice(
        facturas / "2025_03_10_INV000001.xlsx", rows=[_row(1, 1001)]
    )
    make_seitrans_invoice(
        facturas / "2025_04_15_INV000002.xlsx", rows=[_row(1, 1002)]
    )
    make_seitrans_invoice(
        facturas / "2025_04_20_INV000003.xlsx", rows=[_row(1, 1003)]
    )

    result = runner.invoke(
        app,
        [
            "ingest",
            "seitrans",
            "--folder",
            str(tmp_path / "Facturas"),
            "--workbook",
            str(workbook),
        ],
    )
    assert result.exit_code == EXIT_OK, result.output
    assert _datos_row_count(workbook) == 2


def test_auto_discovers_latest_month(tmp_path, default_seur_row):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    facturas = tmp_path / "Facturas"
    make_seur_invoice(
        facturas / "2025" / "03 - Marzo" / "0289992025D0000001.xlsx",
        rows=[default_seur_row(1)],
    )
    make_seur_invoice(
        facturas / "2025" / "04 - Abril" / "0289992025D0000002.xlsx",
        rows=[default_seur_row(2)],
    )

    result = runner.invoke(
        app,
        [
            "ingest",
            "seur",
            "--folder",
            str(facturas),
            "--workbook",
            str(workbook),
        ],
    )
    assert result.exit_code == EXIT_OK, result.output
    assert _datos_row_count(workbook) == 1
    assert "ingested" in result.output


def test_no_files_under_folder_exits_usage(tmp_path):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    result = runner.invoke(
        app,
        [
            "ingest",
            "seur",
            "--folder",
            str(tmp_path / "Facturas"),
            "--workbook",
            str(workbook),
        ],
    )
    assert result.exit_code == EXIT_USAGE


def test_both_file_and_month_is_usage_error(tmp_path, default_seur_row):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    invoice = make_seur_invoice(
        tmp_path / "0289992025D0289264.xlsx", rows=[default_seur_row(1)]
    )
    result = runner.invoke(
        app,
        [
            "ingest",
            "seur",
            "--file",
            str(invoice),
            "--month",
            "2025-04",
            "--workbook",
            str(workbook),
        ],
    )
    assert result.exit_code == EXIT_USAGE


def test_invalid_month_format_exits_usage(tmp_path):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    result = runner.invoke(
        app,
        [
            "ingest",
            "seur",
            "--month",
            "April-2025",
            "--workbook",
            str(workbook),
        ],
    )
    assert result.exit_code == EXIT_USAGE


def test_schema_mismatch_exits_2(tmp_path, default_seur_row):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    bad_columns = list(SEUR_COLUMNS)
    bad_columns[0] = "WRONG"
    invoice = make_seur_invoice(
        tmp_path / "0289992025D0289264.xlsx",
        rows=[default_seur_row(1)],
        columns=tuple(bad_columns),
    )
    result = runner.invoke(
        app,
        ["ingest", "seur", "--file", str(invoice), "--workbook", str(workbook)],
    )
    assert result.exit_code == EXIT_SCHEMA


def test_lock_timeout_exits_3(tmp_path, default_seur_row):
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    invoice = make_seur_invoice(
        tmp_path / "0289992025D0289264.xlsx", rows=[default_seur_row(1)]
    )
    lock_path = workbook.with_suffix(workbook.suffix + LOCK_SUFFIX)
    lock_path.write_text("held by something else")
    try:
        result = runner.invoke(
            app,
            [
                "ingest",
                "seur",
                "--file",
                str(invoice),
                "--workbook",
                str(workbook),
            ],
        )
        assert result.exit_code == EXIT_LOCK
    finally:
        lock_path.unlink(missing_ok=True)


def test_plausibility_failure_exits_5(tmp_path, default_seur_row):
    """Wholesale silent NaN coercion in Peso → CLI exit 5."""
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    rows = []
    for n in range(20):
        r = default_seur_row(n + 1)
        if n >= 5:  # 75% of Pesos become NaN under to_numeric
            r["Peso"] = "1,5"
        rows.append(r)
    invoice = make_seur_invoice(
        tmp_path / "0289992025D0289264.xlsx", rows=rows
    )
    result = runner.invoke(
        app,
        ["ingest", "seur", "--file", str(invoice), "--workbook", str(workbook)],
    )
    assert result.exit_code == EXIT_PLAUSIBILITY
    assert "Peso" in result.output
    assert _datos_row_count(workbook) == 0


def test_manifest_conflict_exits_4(tmp_path, default_seur_row):
    """Two different file contents for the same invoice number → exit 4."""
    workbook = make_empty_seur_workbook(tmp_path / "wb.xlsx")
    row_a = default_seur_row(1)
    row_b = default_seur_row(1)
    row_b["Importe facturado (sin impuestos)"] = 99.99  # forces a different hash

    invoice_a = make_seur_invoice(
        tmp_path / "a" / "0289992025D0289264.xlsx", rows=[row_a]
    )
    invoice_b = make_seur_invoice(
        tmp_path / "b" / "0289992025D0289264.xlsx", rows=[row_b]
    )

    first = runner.invoke(
        app,
        ["ingest", "seur", "--file", str(invoice_a), "--workbook", str(workbook)],
    )
    assert first.exit_code == EXIT_OK

    second = runner.invoke(
        app,
        ["ingest", "seur", "--file", str(invoice_b), "--workbook", str(workbook)],
    )
    assert second.exit_code == EXIT_MANIFEST_CONFLICT
    assert "different hash" in second.output
