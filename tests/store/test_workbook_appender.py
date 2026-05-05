"""Tests for the OneDrive-safe workbook appender."""

from __future__ import annotations

import threading
import time
from pathlib import Path

import pytest
from openpyxl import load_workbook

from courier_automation.parsers.base import SchemaMismatch
from courier_automation.parsers.seur import SEUR_COLUMNS, SeurParser
from courier_automation.store.workbook_appender import (
    LOCK_SUFFIX,
    WorkbookAppender,
    WorkbookLocked,
)


def _datos_row_count(workbook_path: Path) -> int:
    wb = load_workbook(workbook_path, read_only=True)
    try:
        ws = wb["Datos"]
        return ws.max_row - 1  # minus header
    finally:
        wb.close()


def test_append_to_empty_workbook(empty_seur_workbook, seur_invoice_factory):
    invoice_path = seur_invoice_factory()
    parsed = SeurParser().parse(invoice_path)

    appender = WorkbookAppender()
    written = appender.append(
        workbook_path=empty_seur_workbook,
        rows=parsed.rows,
        expected_columns=SEUR_COLUMNS,
    )

    assert written == 1
    assert _datos_row_count(empty_seur_workbook) == 1


def test_append_preserves_existing_rows(
    empty_seur_workbook, seur_invoice_factory, default_seur_row
):
    """Append twice with two different rows; the second append must not erase
    the first."""
    appender = WorkbookAppender()

    first_invoice = seur_invoice_factory(
        rows=[default_seur_row(1)], filename="0289992025D0000001.xlsx"
    )
    second_invoice = seur_invoice_factory(
        rows=[default_seur_row(2), default_seur_row(3)],
        filename="0289992025D0000002.xlsx",
    )

    appender.append(
        workbook_path=empty_seur_workbook,
        rows=SeurParser().parse(first_invoice).rows,
        expected_columns=SEUR_COLUMNS,
    )
    appender.append(
        workbook_path=empty_seur_workbook,
        rows=SeurParser().parse(second_invoice).rows,
        expected_columns=SEUR_COLUMNS,
    )

    assert _datos_row_count(empty_seur_workbook) == 3


def test_append_validates_header_match(
    tmp_path, seur_invoice_factory
):
    """A workbook with the wrong headers must raise SchemaMismatch and stay
    untouched."""
    from openpyxl import Workbook

    bad_workbook = tmp_path / "bad.xlsx"
    wb = Workbook()
    wb.active.title = "Datos"
    bad_headers = list(SEUR_COLUMNS)
    bad_headers[0] = "WRONG"
    wb.active.append(bad_headers)
    wb.save(bad_workbook)
    mtime_before = bad_workbook.stat().st_mtime_ns

    parsed = SeurParser().parse(seur_invoice_factory())
    appender = WorkbookAppender()
    with pytest.raises(SchemaMismatch):
        appender.append(
            workbook_path=bad_workbook,
            rows=parsed.rows,
            expected_columns=SEUR_COLUMNS,
        )
    # File was opened but never written through to (save not reached).
    assert bad_workbook.stat().st_mtime_ns == mtime_before


def test_append_raises_when_datos_sheet_missing(tmp_path, seur_invoice_factory):
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "OtherSheet"
    wb.active.append(list(SEUR_COLUMNS))
    workbook_path = tmp_path / "no-datos.xlsx"
    wb.save(workbook_path)

    parsed = SeurParser().parse(seur_invoice_factory())
    with pytest.raises(SchemaMismatch):
        WorkbookAppender().append(
            workbook_path=workbook_path,
            rows=parsed.rows,
            expected_columns=SEUR_COLUMNS,
        )


def test_lock_retry_then_succeeds(empty_seur_workbook, seur_invoice_factory):
    """Hold the lock briefly in another thread; the appender retries and wins."""
    lock_path = empty_seur_workbook.with_suffix(empty_seur_workbook.suffix + LOCK_SUFFIX)

    held = threading.Event()
    release = threading.Event()

    def hold_lock():
        with open(lock_path, "x") as f:
            f.write("test")
        held.set()
        release.wait(timeout=5)
        lock_path.unlink(missing_ok=True)

    holder = threading.Thread(target=hold_lock)
    holder.start()
    held.wait(timeout=2)

    parsed = SeurParser().parse(seur_invoice_factory())
    appender = WorkbookAppender(lock_retries=10, lock_retry_seconds=0.1)

    def release_after_short_delay():
        time.sleep(0.25)
        release.set()

    threading.Thread(target=release_after_short_delay, daemon=True).start()
    written = appender.append(
        workbook_path=empty_seur_workbook,
        rows=parsed.rows,
        expected_columns=SEUR_COLUMNS,
    )
    holder.join(timeout=2)

    assert written == 1
    assert not lock_path.exists()


def test_lock_timeout_raises_workbook_locked(empty_seur_workbook, seur_invoice_factory):
    lock_path = empty_seur_workbook.with_suffix(empty_seur_workbook.suffix + LOCK_SUFFIX)
    lock_path.write_text("held by something else")

    parsed = SeurParser().parse(seur_invoice_factory())
    appender = WorkbookAppender(lock_retries=2, lock_retry_seconds=0.05)

    with pytest.raises(WorkbookLocked):
        appender.append(
            workbook_path=empty_seur_workbook,
            rows=parsed.rows,
            expected_columns=SEUR_COLUMNS,
        )
    # We must NOT have removed the lock that wasn't ours.
    assert lock_path.exists()
    assert _datos_row_count(empty_seur_workbook) == 0
    lock_path.unlink()


def test_lock_released_after_successful_append(
    empty_seur_workbook, seur_invoice_factory
):
    lock_path = empty_seur_workbook.with_suffix(empty_seur_workbook.suffix + LOCK_SUFFIX)
    parsed = SeurParser().parse(seur_invoice_factory())
    WorkbookAppender().append(
        workbook_path=empty_seur_workbook,
        rows=parsed.rows,
        expected_columns=SEUR_COLUMNS,
    )
    assert not lock_path.exists()


def test_appended_values_round_trip(
    empty_seur_workbook, seur_invoice_factory, default_seur_row
):
    """Specific cells round-trip: postcodes preserve leading zeros, numerics
    survive, dates survive."""
    row = default_seur_row(1)
    row["C. Postal Remitente"] = "08001"
    row["Bultos"] = 7
    row["Peso"] = 12.34

    parsed = SeurParser().parse(
        seur_invoice_factory(rows=[row], filename="0289992025D0289264.xlsx")
    )
    WorkbookAppender().append(
        workbook_path=empty_seur_workbook,
        rows=parsed.rows,
        expected_columns=SEUR_COLUMNS,
    )

    wb = load_workbook(empty_seur_workbook)
    try:
        ws = wb["Datos"]
        headers = [cell.value for cell in ws[1]]
        record = dict(zip(headers, [cell.value for cell in ws[2]]))
        assert record["C. Postal Remitente"] == "08001"
        assert record["Bultos"] == 7
        assert record["Peso"] == pytest.approx(12.34)
        # Fecha Factura was 2025-04-30 in the default row
        assert record["Fecha Factura"].date().isoformat() == "2025-04-30"
    finally:
        wb.close()
