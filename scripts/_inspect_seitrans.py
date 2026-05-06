"""Throwaway: inspect a real Seitrans invoice to learn DOCUMENTO_NUMERO format
and sheet name."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
samples = [
    ROOT / "Operations - Couriers" / "04. Seitrans" / "Facturas" / "2025" / "2025_01_31 3065.xlsx",
    ROOT / "Operations - Couriers" / "04. Seitrans" / "Facturas" / "2025" / "2025_06_30_24633.xlsx",
    ROOT / "Operations - Couriers" / "04. Seitrans" / "Facturas" / "2024" / "2024_12_31 Factura 48172.xlsx",
]

for path in samples:
    if not path.exists():
        print(f"\n--- {path.name} (NOT FOUND) ---")
        continue
    print(f"\n--- {path.name} ---")
    wb = load_workbook(path, read_only=True)
    print(f"  sheets: {wb.sheetnames}")
    wb.close()

    df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
    print(f"  rows: {len(df)}, cols: {len(df.columns)}")
    print(f"  first 3 columns: {list(df.columns[:3])}")
    if "DOCUMENTO_NUMERO" in df.columns:
        print(f"  DOCUMENTO_NUMERO values: {df['DOCUMENTO_NUMERO'].unique()[:3].tolist()}")
        print(f"  DOCUMENTO_NUMERO dtype: {df['DOCUMENTO_NUMERO'].dtype}")
    if "DOCUMENTO_DATA" in df.columns:
        print(f"  DOCUMENTO_DATA sample: {df['DOCUMENTO_DATA'].iloc[0]}")
