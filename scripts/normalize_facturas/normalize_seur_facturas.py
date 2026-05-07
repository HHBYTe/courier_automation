"""Normalize Seur Facturas/<YYYY>/ flat layouts into <YYYY>/<MM> - <Mes>/ subfolders.

For each .xlsx directly under a year folder, read `Fecha Factura` from the
first data row to decide which month subfolder it belongs in. Move the .xlsx
and any same-stem companion file (e.g. .pdf) together.

Dry-run by default. Pass --apply to actually move files.
"""

from __future__ import annotations

import argparse
import datetime as dt
import random
import shutil
import sys
from pathlib import Path

import openpyxl

from _duplicates import dedupe

SAMPLE_SIZE = 3

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_FACTURAS = ROOT / "Operations - Couriers" / "01. Seur" / "Facturas"

SPANISH_MONTHS: dict[int, str] = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def read_invoice_month(path: Path) -> tuple[int, int] | None:
    """Sample SAMPLE_SIZE random data rows and return (year, month) only if all
    samples agree on the same `Fecha Factura` month."""
    # NOTE: do not use read_only=True. Some Seur invoices contain stale
    # `[trash]/` zip entries and a stub `<dimension>` tag; read_only honours
    # the stub and silently reports 1 column / 0 rows. Full load is correct.
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  ! cannot open {path.name}: {e}", file=sys.stderr)
        return None
    try:
        if "Sheet1" not in wb.sheetnames:
            print(f"  ! {path.name}: no Sheet1", file=sys.stderr)
            return None
        ws = wb["Sheet1"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "Fecha Factura" not in headers:
            print(f"  ! {path.name}: no 'Fecha Factura' column", file=sys.stderr)
            return None
        idx = headers.index("Fecha Factura")

        # Read all `Fecha Factura` values once (read_only iterates forward only,
        # and these invoices are small enough that loading one column is cheap).
        values = [
            row[idx] for row in ws.iter_rows(min_row=2, values_only=True)
            if row and row[idx] is not None
        ]
    finally:
        wb.close()

    if not values:
        print(f"  ! {path.name}: no data rows", file=sys.stderr)
        return None

    sample_n = min(SAMPLE_SIZE, len(values))
    sample = random.sample(values, sample_n)

    months: set[tuple[int, int]] = set()
    for v in sample:
        if not isinstance(v, dt.datetime):
            print(f"  ! {path.name}: Fecha Factura sample {v!r} not a date", file=sys.stderr)
            return None
        months.add((v.year, v.month))

    if len(months) != 1:
        print(f"  ! {path.name}: sampled months disagree {sorted(months)}", file=sys.stderr)
        return None
    return months.pop()


def move_pair(xlsx: Path, dest_dir: Path, *, apply: bool) -> list[Path]:
    """Move xlsx and any same-stem companions (e.g. .pdf) into dest_dir."""
    moved: list[Path] = []
    companions = [p for p in xlsx.parent.iterdir()
                  if p.is_file() and p.stem == xlsx.stem]
    for src in companions:
        dst = dest_dir / src.name
        if dst.exists():
            print(f"  ! collision, skipping: {dst}", file=sys.stderr)
            continue
        if apply:
            dest_dir.mkdir(parents=True, exist_ok=True)
            shutil.move(str(src), str(dst))
        moved.append(dst)
    return moved


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=DEFAULT_FACTURAS,
                        help=f"Seur Facturas root (default: {DEFAULT_FACTURAS}).")
    parser.add_argument("--apply", action="store_true",
                        help="Actually move files. Without this, just print the plan.")
    args = parser.parse_args()

    if not args.root.exists():
        print(f"error: {args.root} does not exist", file=sys.stderr)
        return 1

    total_moved = 0
    total_skipped = 0
    for year_dir in sorted(args.root.iterdir()):
        if not (year_dir.is_dir() and year_dir.name.isdigit()):
            continue
        flat_xlsx = sorted(p for p in year_dir.glob("*.xlsx") if p.is_file())
        if not flat_xlsx:
            continue
        print(f"\n[{year_dir.name}] {len(flat_xlsx)} flat .xlsx files")
        for xlsx in flat_xlsx:
            ym = read_invoice_month(xlsx)
            if ym is None:
                total_skipped += 1
                continue
            year, month = ym
            if str(year) != year_dir.name:
                print(f"  ! {xlsx.name}: Fecha Factura year {year} != folder {year_dir.name}, skipping",
                      file=sys.stderr)
                total_skipped += 1
                continue
            dest_dir = year_dir / f"{month:02d} - {SPANISH_MONTHS[month]}"
            moved = move_pair(xlsx, dest_dir, apply=args.apply)
            for m in moved:
                print(f"  {'MOVE' if args.apply else 'PLAN'} {xlsx.parent.name}/{m.name}"
                      f" -> {dest_dir.relative_to(year_dir.parent)}/")
            total_moved += len(moved)

    print(f"\n{'Moved' if args.apply else 'Would move'}: {total_moved} files; "
          f"skipped: {total_skipped}")
    dedupe(args.root, apply=args.apply)
    if not args.apply:
        print("(dry-run) re-run with --apply to actually move and dedupe")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
