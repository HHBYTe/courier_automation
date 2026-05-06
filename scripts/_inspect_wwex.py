"""Throwaway: inspect Wwex schemas across the three file formats AND the
historical workbook (which uses a different 44-col schema)."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

# Pick one file of each extension
samples = {
    "xlsx": ROOT / "Operations - Couriers" / "11. Wwex (US)" / "2025"
    / "2025_05_31 shipment_detail_report.xlsx",
    "xls": ROOT / "Operations - Couriers" / "11. Wwex (US)" / "2025"
    / "2025_01_31 shipment_detail_report.xls",
    "csv": ROOT / "Operations - Couriers" / "11. Wwex (US)" / "2025"
    / "2025_04_30 shipment_detail_report.csv",
}

for ext, path in samples.items():
    print(f"\n--- {ext.upper()} sample: {path.name} ---")
    if not path.exists():
        print(f"  (not found)")
        continue
    if ext == "csv":
        df = pd.read_csv(path, nrows=2)
        print(f"  rows: {len(df)}, cols: {len(df.columns)}")
    elif ext == "xls":
        try:
            wb_sheets = pd.ExcelFile(path, engine="xlrd").sheet_names
            print(f"  sheets: {wb_sheets}")
            df = pd.read_excel(path, sheet_name=0, engine="xlrd", nrows=2)
            print(f"  rows: {len(df)}, cols: {len(df.columns)}")
        except Exception as e:
            print(f"  xlrd error (likely missing): {e}")
            continue
    else:  # xlsx
        wb = load_workbook(path, read_only=True)
        print(f"  sheets: {wb.sheetnames}")
        wb.close()
        df = pd.read_excel(path, sheet_name=0, engine="openpyxl", nrows=2)
        print(f"  rows: {len(df)}, cols: {len(df.columns)}")
    print("  columns:")
    for i, c in enumerate(df.columns, 1):
        print(f"    {i:2d}  {c!r}")

# Historical
hist = ROOT / "Operations - Couriers" / "11. Wwex (US)" / "Wwex USA Shippings Report.xlsx"
print(f"\n--- historical: {hist.name} ---")
wb = load_workbook(hist, read_only=True)
print(f"  sheets: {wb.sheetnames}")
wb.close()
for sn in ["Data", "Datos", "Sheet1"]:
    try:
        h = pd.read_excel(hist, sheet_name=sn, engine="openpyxl", nrows=2)
        print(f"  sheet {sn!r}: shape {h.shape}")
        for i, c in enumerate(h.columns, 1):
            print(f"    {i:2d}  {c!r}")
        break
    except Exception as e:
        print(f"  sheet {sn!r}: {e}")
