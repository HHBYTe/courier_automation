"""Throwaway: confirm UPS CSV schema and historical workbook columns."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

# Pick a sample CSV invoice
sample = next(
    (ROOT / "Operations - Couriers" / "07. UPS (UK)" / "Invoices" / "2025" / "01 - January").glob("*.csv")
)
print(f"--- raw CSV: {sample.name} ---")
df = pd.read_csv(sample, nrows=3, encoding="utf-8", on_bad_lines="warn")
print(f"  rows: 3, cols: {len(df.columns)}")
print("  first 12 columns:")
for i, c in enumerate(df.columns[:12], 1):
    print(f"    {i:3d}  {c!r}")
print("  last 5 columns:")
for i, c in enumerate(df.columns[-5:], len(df.columns) - 4):
    print(f"    {i:3d}  {c!r}")

# Historical workbook
hist = ROOT / "Operations - Couriers" / "07. UPS (UK)" / "UPS Shippings Report.xlsx"
print(f"\n--- historical: {hist.name} ---")
wb = load_workbook(hist, read_only=True)
print(f"  sheets: {wb.sheetnames}")
wb.close()

# Most likely sheet name is `Data`
for sn in ["Data", "Datos", wb.sheetnames[0]]:
    try:
        h = pd.read_excel(hist, sheet_name=sn, engine="openpyxl", nrows=2)
        print(f"  sheet {sn!r}: {h.shape}")
        print(f"    first 5 cols: {list(h.columns[:5])}")
        print(f"    last 5 cols: {list(h.columns[-5:])}")
        break
    except Exception as e:  # noqa: BLE001
        print(f"  sheet {sn!r}: not readable ({e})")

print(f"\n--- raw col count vs historical col count ---")
print(f"  raw CSV: {len(df.columns)}")
