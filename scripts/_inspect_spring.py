"""Throwaway: inspect Spring REPORT and INVOICES schemas."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sample = (
    ROOT / "Operations - Couriers" / "13.Spring (FR)" / "2025"
    / "E2509827_ES_Details of Invoice_O_110003790_2511251351.XLSX"
)

print(f"--- {sample.name} ---")
wb = load_workbook(sample, read_only=True)
print(f"  sheets: {wb.sheetnames}")
wb.close()

for sn in wb.sheetnames:
    df = pd.read_excel(sample, sheet_name=sn, engine="openpyxl", nrows=2)
    print(f"\n  sheet {sn!r}: shape {df.shape}")
    if df.shape[1] <= 30:
        for i, c in enumerate(df.columns, 1):
            print(f"    {i:2d}  {c!r}")
    else:
        print(f"    first 6: {list(df.columns[:6])}")
        print(f"    last 6:  {list(df.columns[-6:])}")
