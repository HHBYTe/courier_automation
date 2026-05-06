"""Throwaway: confirm Correos Express raw schema (header band on row 1)
and historical Datos schema."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sample_raw = (
    ROOT
    / "Operations - Couriers"
    / "05. Correos Express"
    / "Facturas"
    / "2025"
    / "2025_01_31 FAC_UNICO_F2501_14307.xlsx"
)
historical = (
    ROOT / "Operations - Couriers" / "05. Correos Express"
    / "Análisis Envíos Correos Express V2.xlsx"
)

print(f"--- raw: {sample_raw.name} ---")
wb = load_workbook(sample_raw, read_only=True)
print(f"  sheets: {wb.sheetnames}")
wb.close()

raw = pd.read_excel(sample_raw, sheet_name=0, engine="openpyxl", header=None, nrows=5)
print(f"  shape (header=None): {raw.shape}")
print("  first 3 rows (truncated to 8 cols):")
for i, row in raw.head(3).iterrows():
    vals = [str(v)[:25] for v in row.iloc[:8].tolist()]
    print(f"    row {i}: {vals}")

# Now read it as if row 1 is the header
raw_h1 = pd.read_excel(sample_raw, sheet_name=0, engine="openpyxl", header=1)
print(f"\n  shape with header=1: {raw_h1.shape}")
print(f"  columns: {list(raw_h1.columns)}")

print(f"\n--- historical: {historical.name} ---")
wb = load_workbook(historical, read_only=True)
print(f"  sheets: {wb.sheetnames}")
wb.close()

datos = pd.read_excel(historical, sheet_name="Datos", engine="openpyxl", nrows=2)
print(f"  Datos shape: {datos.shape}")
print(f"  columns ({len(datos.columns)}):")
for i, c in enumerate(datos.columns, 1):
    print(f"    {i:2d}  {c!r}")
