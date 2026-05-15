"""Append parser rows to a courier's historical workbook.

The OneDrive-safe write strategy (sidecar lock + off-tree working copy
+ atomic replace) is owned by `LocalStorage.update_xlsx_atomically` in
`courier_automation/storage/local.py`. This module is now a thin shell:
it provides the openpyxl row-append mutator and delegates the
transactional dance to the injected `Storage`.

When no `Storage` is passed, `append()` builds a `LocalStorage` rooted
at the workbook's parent directory on the fly — preserving the legacy
`WorkbookAppender().append(workbook_path=...)` calling convention.
Later refactor steps inject an explicit Storage at the call sites.
"""

from __future__ import annotations

import datetime as _dt
import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from courier_automation.parsers.base import SchemaMismatch, assert_schema
from courier_automation.storage import LocalStorage, OpsLocator, Storage
from courier_automation.storage.base import StorageLocked
from courier_automation.storage.local import (  # re-exported for back-compat
    LOCK_SUFFIX,  # noqa: F401
)

log = logging.getLogger(__name__)

DATOS_SHEET = "Datos"

# Backward-compatible alias for callers importing `WorkbookLocked` from
# this module. The exception class itself now lives in storage.base.
WorkbookLocked = StorageLocked

__all__ = [
    "DATOS_SHEET",
    "WorkbookAppender",
    "WorkbookLocked",
    "export_rows",
    "export_parquet",
]


class WorkbookAppender:
    """Append rows to a master workbook's data sheet, atomically.

    Holds the per-mutation policy (which sheet to write, how many
    retries to make on a held lock). The transactional `read → mutate
    → publish` cycle lives in the `Storage` backend.
    """

    def __init__(
        self,
        *,
        sheet_name: str = DATOS_SHEET,
        lock_retries: int = 6,
        lock_retry_seconds: float = 5.0,
        working_dir: Path | None = None,
        storage: Storage | None = None,
    ) -> None:
        self.sheet_name = sheet_name
        self.lock_retries = lock_retries
        self.lock_retry_seconds = lock_retry_seconds
        self.working_dir = working_dir
        self.storage = storage

    def append(
        self,
        *,
        workbook_path: Path,
        rows: pd.DataFrame,
        expected_columns: tuple[str, ...],
    ) -> int:
        """Append `rows` to the workbook's data sheet. Returns rows written."""
        workbook_path = Path(workbook_path)
        if not workbook_path.exists():
            raise FileNotFoundError(f"workbook not found: {workbook_path}")

        # Resolve `workbook_path` to an `OpsLocator` against the injected
        # storage. Three cases:
        #   1. No storage injected → build a scoped LocalStorage at the
        #      workbook's parent dir (legacy WorkbookAppender behaviour).
        #   2. LocalStorage injected and workbook is under its ops_root
        #      → compute the relative locator.
        #   3. LocalStorage injected but workbook is OUTSIDE ops_root
        #      (e.g. CLI test passes an absolute tmp_path workbook with a
        #      repo-rooted storage) → fall back to case 1.
        storage = self.storage
        loc: OpsLocator | None = None
        if isinstance(storage, LocalStorage):
            try:
                rel = workbook_path.resolve().relative_to(
                    storage.ops_root.resolve()
                )
                loc = OpsLocator(rel.as_posix())
            except ValueError:
                storage = None  # fall through to the scoped-local branch
        elif storage is not None:
            # Non-local backend: caller is responsible for the locator
            # mapping. Use the bare filename as a placeholder.
            loc = OpsLocator(workbook_path.name)

        if storage is None:
            storage = LocalStorage(
                ops_root=workbook_path.parent,
                working_dir=self.working_dir,
            )
            loc = OpsLocator(workbook_path.name)
        assert loc is not None  # one of the branches above set it

        def _mutate(local_copy: Path) -> int:
            return self._append_to_workbook(local_copy, rows, expected_columns)

        written = storage.update_xlsx_atomically(
            loc,
            _mutate,
            retries=self.lock_retries,
            retry_seconds=self.lock_retry_seconds,
        )
        log.info(
            "workbook append: %s += %d rows (sheet=%s)",
            workbook_path.name,
            written,
            self.sheet_name,
        )
        return written

    def _append_to_workbook(
        self,
        path: Path,
        rows: pd.DataFrame,
        expected_columns: tuple[str, ...],
    ) -> int:
        wb = load_workbook(path, keep_vba=True)
        try:
            if self.sheet_name not in wb.sheetnames:
                raise SchemaMismatch(
                    f"workbook is missing the {self.sheet_name!r} sheet "
                    f"(found: {wb.sheetnames})"
                )
            ws = wb[self.sheet_name]
            actual_headers = tuple(
                cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
            )
            assert_schema(
                pd.DataFrame(columns=list(actual_headers)), expected_columns
            )

            written = 0
            for record in rows.to_dict(orient="records"):
                ws.append(
                    [_to_excel_value(record[col]) for col in expected_columns]
                )
                written += 1
            wb.save(path)
            return written
        finally:
            wb.close()


def export_rows(
    *,
    output_path: Path,
    rows: pd.DataFrame,
    expected_columns: tuple[str, ...],
    sheet_name: str = DATOS_SHEET,
    numeric_columns: tuple[str, ...] = (),
    date_formats: dict[str, str] | None = None,
    number_formats: dict[str, str] | None = None,
) -> int:
    """Write `rows` to a fresh xlsx with `expected_columns` as the header row.

    Use this to stage append-ready output without loading the master workbook.
    The caller is responsible for ensuring the master's schema matches
    `expected_columns` (this helper does no live header check).
    """
    if output_path.exists():
        raise FileExistsError(
            f"export target already exists: {output_path} "
            f"(deal with the prior export before producing a new one)"
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    numeric_set = set(numeric_columns)
    fmt_overrides = date_formats or {}
    num_fmt_overrides = number_formats or {}
    # Map column index (1-based) → number format. Columns not listed get the
    # default dd/mm/yyyy when the cell value is a date/datetime.
    column_format_by_idx = {
        i + 1: fmt_overrides[col]
        for i, col in enumerate(expected_columns)
        if col in fmt_overrides
    }
    numeric_format_by_idx = {
        i + 1: num_fmt_overrides[col]
        for i, col in enumerate(expected_columns)
        if col in num_fmt_overrides
    }
    wb = Workbook()
    try:
        ws = wb.active
        ws.title = sheet_name
        ws.append(list(expected_columns))
        written = 0
        for record in rows.to_dict(orient="records"):
            ws.append([
                _to_numeric_or_passthrough(record[col]) if col in numeric_set
                else _to_excel_value(record[col])
                for col in expected_columns
            ])
            written += 1
        # Apply per-column number-format overrides. Date cells default to
        # the master's Spanish dd/mm/yyyy; numeric cells without an override
        # are left at General (Excel's default).
        for row in ws.iter_rows(min_row=2, max_row=written + 1):
            for cell in row:
                if isinstance(cell.value, (_dt.date, _dt.datetime)):
                    cell.number_format = column_format_by_idx.get(
                        cell.column, "dd/mm/yyyy"
                    )
                elif isinstance(cell.value, (int, float)):
                    fmt = numeric_format_by_idx.get(cell.column)
                    if fmt is not None:
                        cell.number_format = fmt
        wb.save(output_path)
        return written
    finally:
        wb.close()


def export_parquet(
    *,
    output_path: Path,
    rows: pd.DataFrame,
    expected_columns: tuple[str, ...],
) -> int:
    """Write `rows` to a parquet file with `expected_columns` as the schema.

    Parallel-write companion to `export_rows`: same input, parquet output.
    Snappy compression, no index. Object-dtype columns (UPS' mixed
    int+text Charge Description Code is the live case) are coerced to
    pandas StringDtype so pyarrow gets a clean type per column.
    """
    if output_path.exists():
        raise FileExistsError(
            f"parquet target already exists: {output_path} "
            f"(deal with the prior export before producing a new one)"
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df = rows.loc[:, list(expected_columns)].copy()
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].map(
                lambda v: None
                if v is None or (isinstance(v, float) and pd.isna(v))
                else str(v)
            ).astype("string")
    df.to_parquet(output_path, engine="pyarrow", compression="snappy", index=False)
    return len(df)


def _to_numeric_or_passthrough(val: object) -> object:
    """For export numeric columns: turn an all-digit string into int.
    Leave anything else (None, NaN, embedded letters, leading zeros) as-is so
    we don't silently drop information."""
    cleaned = _to_excel_value(val)
    if isinstance(cleaned, str) and cleaned.isdigit() and not (
        len(cleaned) > 1 and cleaned.startswith("0")
    ):
        return int(cleaned)
    return cleaned


def _to_excel_value(val: object) -> object:
    """Convert a pandas/numpy value into something openpyxl writes cleanly.

    NaN/NaT/<NA> → None; pandas Timestamp → datetime; pandas Int64/Float64 NA → None.
    """
    if val is None:
        return None
    if isinstance(val, pd.Timestamp):
        if pd.isna(val):
            return None
        py = val.to_pydatetime()
        if py.hour == 0 and py.minute == 0 and py.second == 0 and py.microsecond == 0:
            return py.date()
        return py
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        # pd.isna can choke on some array-like values; we'll just pass through.
        pass
    return val
